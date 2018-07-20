#coding=utf-8
import os
import openpyxl
import glob
import re
from openpyxl.styles import PatternFill,Border,Side,Alignment,Protection,Font
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles import Font

alignC = Alignment()
alignC.horz = Alignment(horizontal='general')
alignC.vert = Alignment(vertical='bottom')

alignL = Alignment()
alignL.horz = Alignment(horizontal='left')
alignL.vert = Alignment(vertical='bottom')

alignR = Alignment()
alignR.horz = Alignment(horizontal='right')
alignR.vert = Alignment(vertical='bottom')

dic_Alignment = {'left': alignL, 'center': alignC, 'right': alignR}
dic_Font = {}  # 字体
dic_Border = {}  # 表格边框
dic_Style = {}  # 单元格样式
dic_RowStyle = {}  # 行高
dic_season = {'S':'夏秋',
              'W':'冬春'}

def GetAlignment():
    
    alignment = Alignment(horizontal = 'center' , vertical = 'center')   
    return alignment



def GetBorder():
    border = Border(left=Side(border_style='thin',color='000000'),
                    right=Side(border_style='thin',color='000000'),
                    top=Side(border_style='thin',color='000000'),
                    bottom=Side(border_style='thin',color='000000'))
    return border

def GetFont():
    font = Font(name='Calibri',
                size = 12,
                bold = False,
                italic = False,
                vertAlign = 'baseline',
                strike=False,
                color='000000')
    return font

class CFP_737:

    def Get_Info(self,txt,filename):
        header_re = "DST\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)"
        m = re.search(header_re,txt,re.M)
        self.DEST,self.POAFUEL,self.LEGTIME,self.DIST,self.ETA,self.MTOW,self.LDW,self.ZFW,self.PL= m.groups()
        self.TTLFUEL = re.search('TTL\s*(\S+)',txt).group(1)
        self.WIND = re.search('WIND\s*(\S+)',txt).group(1)
        self.ALTN = re.search('ALT\s*(\S+)',txt).group(1)
        self.DEP = filename[-12:10]
        self.DEST = filename[-9:-6]
        self.SEASON = filename[-5]
        self.AC_TYPE = 'B737F'

    def write_sheet(self,ws):

        row_max = ws.max_row
        col_max = ws.max_column

        DATA = [self.AC_TYPE,
                self.DEP+ self.DEST,
                self.DEP,
                self.DEST,
                dic_season[self.SEASON],
                self.DIST,
                int(self.TTLFUEL),
                int(int(self.TTLFUEL)*2.2046),
                int(self.POAFUEL),
                int(int(self.POAFUEL)*2.2046),
                int(self.MTOW),
                int(int(self.MTOW)*2.2046),
                self.LEGTIME,
                int(self.PL),
                int(int(self.PL)*2.2046),
                self.ALTN,
                self.WIND,
                ' '
                ]
        
        for row in range(row_max+1,row_max+2):
            for col in range(0,len(DATA)):
                cell = ws.cell(column = col+1, row = row, value = DATA[col])
                cell.font = GetFont()
                cell.border = GetBorder()
                cell.alignment = GetAlignment()

        print(row_max)
                    
class CFP_747:

    def Get_Info(self,txt,filename):
        header_re = "POA\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)\s*(\S+)"
        m = re.search(header_re,txt,re.M)
        self.DEST,self.POAFUEL,self.LEGTIME,self.NAM,self.NGM,self.MTOW,self.LDW,self.ZFW,self.PL= m.groups()
        self.RAMPFUEL = re.search('RAMP\s*(\S+)',txt).group(1)
        self.WIND = re.search('WIND\s*(\S+)',txt).group(1)
        self.ALTN = re.search('ALT\s*(\S+)',txt).group(1)
        self.DEP = filename[-12:10]
        self.DEST = filename[-9:-6]
        self.SEASON = filename[-5]
        self.AC_TYPE = 'B747F'

    def write_sheet(self,ws):

        row_max = ws.max_row
        col_max = ws.max_column

        DATA = [self.AC_TYPE,
                self.DEP+ self.DEST,
                self.DEP,
                self.DEST,
                dic_season[self.SEASON],
                self.NGM,
                int(self.RAMPFUEL),
                int(int(self.RAMPFUEL)*0.4536),
                int(self.POAFUEL),
                int(int(self.POAFUEL)*0.4536),
                int(self.MTOW),
                int(int(self.MTOW)*0.4536),
                self.LEGTIME,
                int(self.PL),
                int(int(self.PL)*0.4536),
                self.ALTN,
                self.WIND,
                ' '
                ]
        
        for row in range(row_max+1,row_max+2):
            for col in range(0,len(DATA)):
                cell = ws.cell(column = col+1, row = row, value = DATA[col])
                cell.font = GetFont()
                cell.border = GetBorder()
                cell.alignment = GetAlignment()

        print(row_max)

if __name__ == '__main__':

    work_book = openpyxl.load_workbook('D:\CFP\天货航线测算.xlsx')
    print('请输入机型:\n1-B737-300F\n2-B737-400F\n3-B747F ')
    AC_TYPE_INPUT = int(input())
    if AC_TYPE_INPUT == 1:
        AC_TYPE = 'B737-300F'
    elif AC_TYPE_INPUT == 2:
        AC_TYPE = 'B737-400F'
    elif AC_TYPE_INPUT == 3:
        AC_TYPE = 'B747F'
    work_sheet = work_book[AC_TYPE]
    
    txt_filenames = glob.glob('d:\\CFP\\*.txt')
    CFP_numbers = 0
    for filename in txt_filenames:
        CFP_numbers += 1        
        txt_file = open(filename, 'r')
        txt = txt_file.read()
        if AC_TYPE_INPUT ==1:
            X = CFP_737()
            X.Get_Info(txt,filename)
            X.write_sheet(work_sheet)
        elif AC_TYPE_INPUT ==2:
            X = CFP_737()
            X.Get_Info(txt,filename)
            X.write_sheet(work_sheet)
        else: 
            X = CFP_747()
            X.Get_Info(txt,filename)
            X.write_sheet(work_sheet)
        txt_file.close()

        
    work_book.save('D:\CFP\天货航线测算.xlsx')


